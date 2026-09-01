#!/usr/bin/env python3
"""
MAS AI Labs — Excel Sprint Tracker Generator & Database Sync
Author: MAS AI PM
Description: Parses the Markdown Sprint files and generates a multi-tab,
             formatted Excel Workbook (MAS_AI_LABS_SPRINT_TRACKER.xlsx)
             and CSV backup (sprint_tasks.csv) to ensure 100% crash-proof structured storage.
"""

import os
import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__)))
OUTPUT_XLSX = os.path.join(BASE_DIR, "MAS_AI_LABS_SPRINT_TRACKER.xlsx")
OUTPUT_CSV = os.path.join(BASE_DIR, "sprint_tasks.csv")

SPRINT_FILES = [
    (1, "Sprint 1 (Week 1: Sept 1–4)", os.path.join(BASE_DIR, "SPRINT_01_WEEK_01.md")),
    (2, "Sprint 2 (Week 2: Sept 7–11)", os.path.join(BASE_DIR, "SPRINT_02_WEEK_02.md")),
    (3, "Sprint 3 (Week 3: Sept 14–18)", os.path.join(BASE_DIR, "SPRINT_03_WEEK_03.md")),
    (4, "Sprint 4 (Week 4: Sept 21–30)", os.path.join(BASE_DIR, "SPRINT_04_WEEK_04.md")),
]

def parse_markdown_sprint_file(file_path):
    tasks = []
    if not os.path.exists(file_path):
        return tasks

    current_compartment = "General"
    task_regex = re.compile(
        r"\|\s*\*\*(S\d+\.\d+)\*\*\s*\|\s*([^|]+)\s*\|\s*\*\*([^*]+)\*\*(?:\s*\*\([^*]+\)\*)?\s*\|\s*([^|]+)\s*\|\s*([^|]+)\s*\|\s*([^|]+)\s*\|\s*([^|]+)\s*\|\s*([^|]+)\s*\|\s*([^|]+)\s*\|\s*([^|]+)\s*\|"
    )

    with open(file_path, "r", encoding="utf-8") as f:
        for line in f:
            if line.startswith("## 📦"):
                current_compartment = line.replace("## 📦", "").strip()
            match = task_regex.search(line)
            if match:
                tasks.append({
                    "compartment": current_compartment,
                    "id": match.group(1).strip(),
                    "task": match.group(2).strip(),
                    "owner": match.group(3).strip(),
                    "date": match.group(4).strip(),
                    "status": match.group(5).strip().replace("`", ""),
                    "expected": match.group(6).strip(),
                    "actual": match.group(7).strip(),
                    "blocker": match.group(8).strip(),
                    "delay_reason": match.group(9).strip(),
                    "rag": match.group(10).strip(),
                })
    return tasks

def create_excel_workbook():
    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid") # Dark Gray / Slate
    comp_header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid") # Blue
    thin_border = Border(
        left=Side(style='thin', color='E5E7EB'),
        right=Side(style='thin', color='E5E7EB'),
        top=Side(style='thin', color='E5E7EB'),
        bottom=Side(style='thin', color='E5E7EB')
    )

    all_tasks_flat = []

    for s_num, tab_name, f_path in SPRINT_FILES:
        tasks = parse_markdown_sprint_file(f_path)
        ws = wb.create_sheet(title=f"Sprint {s_num}")

        headers = [
            "Compartment", "Task ID", "Task / Activity", "Lead Owner",
            "Target Date", "Status", "Expected Outcome", "Actual Outcome",
            "Blocker", "Delay Reason", "RAG"
        ]

        ws.append(headers)
        for col_num, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for t in tasks:
            row_data = [
                t["compartment"], t["id"], t["task"], t["owner"],
                t["date"], t["status"], t["expected"], t["actual"],
                t["blocker"], t["delay_reason"], t["rag"]
            ]
            ws.append(row_data)
            all_tasks_flat.append({"sprint": f"Sprint {s_num}", **t})

            row_idx = ws.max_row
            for c_idx in range(1, len(row_data) + 1):
                c = ws.cell(row=row_idx, column=c_idx)
                c.border = thin_border
                c.alignment = Alignment(vertical="center", wrap_text=True)
                if c_idx in [2, 4, 5, 6, 11]:
                    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Auto-fit column widths
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 45)

    # Master Overview Tab
    ws_master = wb.create_sheet(title="All Sprints Master", index=0)
    master_headers = [
        "Sprint", "Compartment", "Task ID", "Task / Activity", "Lead Owner",
        "Target Date", "Status", "Expected Outcome", "Actual Outcome",
        "Blocker", "Delay Reason", "RAG"
    ]
    ws_master.append(master_headers)
    for col_num in range(1, len(master_headers) + 1):
        cell = ws_master.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for t in all_tasks_flat:
        ws_master.append([
            t["sprint"], t["compartment"], t["id"], t["task"], t["owner"],
            t["date"], t["status"], t["expected"], t["actual"],
            t["blocker"], t["delay_reason"], t["rag"]
        ])
        row_idx = ws_master.max_row
        for c_idx in range(1, len(master_headers) + 1):
            c = ws_master.cell(row=row_idx, column=c_idx)
            c.border = thin_border
            c.alignment = Alignment(vertical="center", wrap_text=True)

    for col in ws_master.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws_master.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 45)

    wb.save(OUTPUT_XLSX)
    print(f"✅ Excel Workbook generated successfully: {OUTPUT_XLSX}")

    # Also save structured CSV backup
    import csv
    with open(OUTPUT_CSV, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=["sprint", "compartment", "id", "task", "owner", "date", "status", "expected", "actual", "blocker", "delay_reason", "rag"])
        writer.writeheader()
        writer.writerows(all_tasks_flat)
    print(f"✅ CSV database backup generated successfully: {OUTPUT_CSV}")

if __name__ == "__main__":
    create_excel_workbook()
